"""
============================
Author:柠檬班-木森
Time:2019/10/24
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
import openpyxl

# 第一步：打开excle文件

workbook = openpyxl.load_workbook("cases.xlsx")

# 第二步：选择工作薄中的某个表单（sheet）
sh = workbook["test"]

# 读取表格的内容
# print(sh.cell(row=1, column=3).value)
#
# # 往表格中写入内容
# sh.cell(row=1,column=4).value = 999
# sh.cell(row=1,column=5,value=8888)
#
# # 写入完之后，一定要保存
# workbook.save("cases.xlsx")

# 获取表单中最大的行
# max_row = sh.max_row
# print(max_row)
#
# # 获取表单中最大的列
# max_column = sh.max_column
# print(max_column)


# 按行获取表单中的，所有格子对象，每一行的格子放在一个元组当中/
rows = sh.rows
for r in rows:
    print(r)

"""
============================
Author:柠檬班-木森
Time:2019/10/24
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
import openpyxl

# 第一步：打开excle文件
workbook = openpyxl.load_workbook("cases.xlsx")

# 第二步：选择工作薄中的某个表单（sheet）
sh = workbook["register"]

# 获取最大行
max_row = sh.max_row

cases = []

for row in range(2,max_row+1):
    # print(row)
    case_data= {}
    case_data["case_id"] = sh.cell(row=row, column=1).value
    case_data["data"] = sh.cell(row=row, column=2).value
    case_data["excepted"] = sh.cell(row=row, column=3).value
    case_data["result"] = sh.cell(row=row, column=4).value
    cases.append(case_data)
    # print(case_data)

print(cases)